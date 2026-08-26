"""Módulo para facilidades no manuseio de recursos comuns no desenvolvimento."""

from typing import Union

__all__ = [
    'abrir_arquivo_em_bytes',
    'abrir_arquivo_excel',
    'abrir_arquivo_pdf',
    'abrir_arquivo_texto',
    'adicionar_ao_zip',
    'alterar_arquivo_texto',
    'caminho_existente',
    'cls',
    'coletar_arvore_caminho',
    'coletar_caminho_absoluto',
    'coletar_extensao_arquivo',
    'coletar_idioma_so',
    'coletar_versao_so',
    'coletar_nome_arquivo',
    'coletar_nome_guias_arquivo_excel',
    'coletar_pid',
    'coletar_tamanho',
    'coletar_versao_arquivo',
    'compactar',
    'converter_pdf_em_imagem',
    'copiar_arquivo',
    'copiar_pasta',
    'criar_arquivo_texto',
    'criar_pasta',
    'descompactar',
    'escrever_em_arquivo',
    'excluir_arquivo',
    'excluir_pasta',
    'executar_comando_terminal',
    'extrair_texto_ocr',
    'finalizar_processo',
    'gravar_log_em_arquivo',
    'janela_dialogo',
    'ler_variavel_ambiente',
    'logar',
    'pasta_esta_vazia',
    'processo_existente',
    'recortar',
    'remover_acentos',
    'renomear',
    'retornar_arquivos_em_pasta',
    'retornar_data_hora_atual',
    'transformar_arquivo_em_base64',
]


def abrir_arquivo_em_bytes(caminho):
    """Lê todo o conteúdo de um arquivo e devolve seus bytes crus.

    Parâmetros:
        caminho: Caminho do arquivo a ser lido, relativo ou absoluto.

    Retorna:
        bytes: Conteúdo integral do arquivo.

    Exceções:
        FileNotFoundError: Quando o caminho informado não existe.

    Exemplos:
        >>> conteudo = abrir_arquivo_em_bytes(
        ...     caminho='logo.png',
        ... )
        >>> conteudo[:4]
        b'\\x89PNG'
    """
    # Importa recursos do módulo Path
    from pathlib import Path

    # abre um arquivo de texto e coleta o conteúdo em bytes
    arquivo = Path(caminho).read_bytes()

    # retorna o conteúdo do arquivo
    return arquivo


def abrir_arquivo_excel(
    arquivo_excel: str,
    guia: str = '',
    manter_macro: bool = True,
    manter_links: bool = True,
):
    """Lê uma planilha do Excel e devolve seu conteúdo como lista de linhas.

    Parâmetros:
        arquivo_excel: Caminho do arquivo Excel a ser aberto. Aceita os
            formatos `.xls`, `.xlsx` e `.xlsm`.
        guia: Nome da aba a ser lida. Aceita o nome de uma guia ou string
            vazia; string vazia usa a aba ativa e, nos `.xls`, a primeira.
        manter_macro: Aceita `True` para preservar o código VBA de arquivos
            `.xlsm` e `False` para não preservá-lo. Ignorado em `.xls`.
        manter_links: Aceita `True` para preservar vínculos com pastas de
            trabalho externas e `False` para não preservá-los. Ignorado em
            `.xls`.

    Retorna:
        list[list]: Uma lista por linha da planilha, cada uma com os
            valores das células na ordem das colunas. Células vazias vêm
            como `None`.

    Exceções:
        FileNotFoundError: Quando o arquivo não existe.
        KeyError: Quando a guia informada não existe na planilha.

    Exemplos:
        >>> tabela = abrir_arquivo_excel(
        ...     arquivo_excel='base.xlsx',
        ...     guia='Clientes',
        ... )
        >>> tabela[0]
        ['Nome', 'CPF', 'Limite']
    """

    # Importa recursos do módulo openpyxl
    import xlrd  # type: ignore
    from openpyxl import load_workbook

    # trata o caminho com o objeto Path
    caminho_excel = coletar_caminho_absoluto(arquivo_excel)
    extensao_arquivo_excel = coletar_extensao_arquivo(caminho_excel)

    # define um valor padrão e inicial à lista
    tabela_excel = []

    if extensao_arquivo_excel[0].upper() == '.XLS':
        # abre um arquivo de Excel e coleta o conteúdo
        conteudo_excel = xlrd.open_workbook(
            caminho_excel,
        )

        # seleciona a guia à trabalhar
        if guia == '':
            aba_ativa = conteudo_excel.sheet_by_index(0)
        else:
            aba_ativa = conteudo_excel.sheet_by_name(guia)

        # para cada linha do conteúdo coletado
        for indice_linha in range(aba_ativa.nrows):
            # adiciona a linha na tabela
            tabela_excel.append(aba_ativa.row_values(indice_linha))
    else:
        # abre um arquivo de Excel e coleta o conteúdo
        conteudo_excel = load_workbook(
            caminho_excel,
            keep_vba=manter_macro,
            keep_links=manter_links,
        )

        # seleciona a guia à trabalhar
        if guia == '':
            aba_ativa = conteudo_excel.active
        else:
            aba_ativa = conteudo_excel[guia]

        # para cada linha do conteúdo coletado
        for linhas in aba_ativa.values:
            # define um valor padrão e inicial à lista
            linha = []

            # para cada valor na célula da linha
            for celula in linhas:
                # adiciona o valor na linha
                linha.append(celula)

            # adiciona a linha na tabela
            tabela_excel.append(linha)

    # retorna o conteúdo da tabela
    return tabela_excel


def abrir_arquivo_pdf(
    arquivo_pdf: str,
    senha_pdf: Union[str | None] = None,
    paginacao: Union[int | tuple[int]] = 0,
    orientacao: int = 0,
):
    """Extrai o texto de um arquivo PDF, página a página.

    Parâmetros:
        arquivo_pdf: Caminho do arquivo PDF a ser lido.
        senha_pdf: Senha do PDF, quando protegido. `None` para PDFs
            sem proteção.
        paginacao: Página ou páginas a extrair, contadas a partir de 1.
            Aceita um `int` ou uma tupla de `int`. Aceita `0` para o
            documento inteiro ou números a partir de `1` para páginas
            específicas.
        orientacao: Orientação considerada durante a extração. Aceita `0`,
            `90`, `180` ou `270` graus.

    Retorna:
        list[list[str]]: Uma lista por página, cada uma contendo as
            linhas de texto daquela página.

    Exceções:
        TypeError: Quando `paginacao` contém itens não numéricos.

    Exemplos:
        >>> paginas = abrir_arquivo_pdf(
        ...     arquivo_pdf='contrato.pdf',
        ...     paginacao=(1, 2),
        ... )
        >>> paginas[0][0]
        'CONTRATO DE PRESTAÇÃO DE SERVIÇOS'
    """

    # Importa recursos do módulo PyPDF2
    from PyPDF2 import PdfReader

    # trata o caminho com o objeto Path
    caminho_pdf = coletar_caminho_absoluto(arquivo_pdf)

    # abre um arquivo de PDF e coleta o conteúdo
    conteudo_pdf = PdfReader(
        stream=caminho_pdf,
        password=senha_pdf,
        strict=False,
    )

    # define um valor padrão e inicial à lista
    lista_paginacao = []

    # caso o tipo do parâmetro 'paginacao' seja do tipo int
    if isinstance(paginacao, int):
        # transforma 'paginacao' em tupla
        paginacao = (paginacao,)

    # se paginacao for igual à 0
    if 0 in paginacao:
        # adiciona todas as páginas para a lista de paginações
        lista_paginacao = conteudo_pdf.pages
    else:
        # para cada valor do índice de 'paginação'
        for indice in paginacao:
            # caso índice seja do tipo int
            if isinstance(indice, int) is True:
                # ajusta o índice ao padrão de índice em listas da linguagem
                indice = indice - 1

                # adiciona a paginação solicitada à lista de paginações
                lista_paginacao.append(conteudo_pdf.getPage(indice))
            # caso índice não seja do tipo int
            else:
                # levanta exceção de tipo incorreto
                raise TypeError(
                    'Parâmetro `paginacao` aceita somente ítens numéricos (int).'
                )

    # define um valor padrão e inicial à lista
    lista_texto_pdf = []

    # para cada página do conteúdo coletado
    for pagina in lista_paginacao:
        # adiciona o valor na linha separando por páginas [n] e linhas [n][n]
        lista_texto_pdf.append(
            pagina.extract_text(orientations=orientacao).splitlines()
        )

    # retorna o conteúdo coletado em lista
    return lista_texto_pdf


def abrir_arquivo_texto(caminho, encoding='utf8'):
    """Lê um arquivo de texto e devolve todo o seu conteúdo em uma string.

    Parâmetros:
        caminho: Caminho do arquivo a ser lido.
        encoding: Codificação usada na leitura. Use 'latin-1' ou
            'cp1252' para arquivos gerados por sistemas legados.

    Retorna:
        str: Conteúdo completo do arquivo, com as quebras de linha
            preservadas.

    Exceções:
        FileNotFoundError: Quando o caminho informado não existe.
        UnicodeDecodeError: Quando o encoding informado não corresponde
            ao do arquivo.

    Exemplos:
        >>> abrir_arquivo_texto(
        ...     caminho='config.ini',
        ... )
        '[padrao]\\nusuario = admin\\n'
    """
    # Importa recursos do módulo Path
    from pathlib import Path

    # abre um arquivo de texto e coleta o conteúdo
    arquivo = Path(caminho).read_text(encoding=encoding)

    # retorna o conteúdo do arquivo
    return arquivo


def adicionar_ao_zip(
    caminho: str,
    arquivo_destino: str,
    recursivo: bool = False,
):
    """Acrescenta arquivos a um `.zip` já existente, sem recriá-lo.

    Parâmetros:
        caminho: Arquivo ou pasta a ser adicionado ao pacote.
        arquivo_destino: Caminho do `.zip` que receberá o conteúdo.
            É criado caso ainda não exista.
        recursivo: Aceita `True` para tratar `caminho` como pasta e
            incluir todo o seu conteúdo, ou `False` para adicionar somente
            o arquivo informado.

    Retorna:
        bool: `True` quando a operação é concluída.

    Exemplos:
        >>> adicionar_ao_zip(
        ...     caminho='relatorio.pdf',
        ...     arquivo_destino='lote.zip',
        ... )
        True
        >>> adicionar_ao_zip(
        ...     caminho='saida',
        ...     arquivo_destino='lote.zip',
        ...     recursivo=True,
        ... )
        True
    """
    # Importa recursos do módulo zipfile
    from zipfile import ZipFile

    if recursivo is True:
        filtro = './**/*'
        lista_caminhos = retornar_arquivos_em_pasta(
            caminho=caminho,
            filtro=filtro,
        )
    else:
        lista_caminhos = [
            caminho,
        ]

    with ZipFile(arquivo_destino, 'a') as objeto_zip:
        for arquivo in lista_caminhos:
            if len(lista_caminhos) == 1:
                caminho = coletar_arvore_caminho(lista_caminhos[0])

            caminho_interno_zip = arquivo.replace(caminho, '')
            objeto_zip.write(
                filename=arquivo,
                arcname=caminho_interno_zip,
            )

    return True


def alterar_arquivo_texto(
    caminho,
    linha_atual,
    linha_alterada,
    multilinhas=False,
    encoding_entrada='utf8',
    encoding_saida='utf8',
):
    """Substitui o conteúdo de linhas de um arquivo de texto no próprio local.

    Parâmetros:
        caminho: Caminho do arquivo a ser alterado.
        linha_atual: Trecho procurado. A linha é considerada candidata
            quando o contém, não sendo necessário casar a linha inteira.
        linha_alterada: Texto que substituirá a linha encontrada.
        multilinhas: Aceita `True` para alterar todas as ocorrências ou
            `False` para alterar somente a primeira.
        encoding_entrada: Codificação usada para ler o arquivo.
        encoding_saida: Codificação usada para regravá-lo.

    Retorna:
        list[str]: As linhas do arquivo já com as alterações aplicadas.

    Exemplos:
        >>> alterar_arquivo_texto(
        ...     caminho='config.ini',
        ...     linha_atual='ambiente=dev',
        ...     linha_alterada='ambiente=prod',
        ... )
        ['[padrao]', 'ambiente=prod']
    """
    # Importa recursos do módulo Path
    from pathlib import Path

    # trata o caminho com o objeto Path
    arquivo = Path(caminho)
    # abre um arquivo de texto e coleta o conteúdo
    conteudo = arquivo.read_text(encoding=encoding_entrada)
    # divide o conteúdo por linhas ('\n')
    conteudo = conteudo.splitlines()

    # define um valor padrão para a variável
    validacao_multilinhas = True

    # para cada linha do arquivo .txt aberto
    for linha_arquivo in range(len(conteudo)):
        # se for a primeira linha, define o modo de escrita
        if linha_arquivo == 0:
            modo = 'w'
        # se não for a primeira linha, define o modo de concatenação
        else:
            modo = 'a'

        # abre um arquivo de texto para alteração
        with open(arquivo, modo, encoding=encoding_saida) as arquivo_aberto:
            # caso seja definido multiplas linhas:
            if validacao_multilinhas is True:
                # se a linha atual corresponder ao conteúdo passado no parâmetro
                if conteudo[linha_arquivo].__contains__(linha_atual):
                    # substitui a linha atual pelo conteúdo passada no parâmetro
                    conteudo[linha_arquivo] = linha_atual.replace(
                        linha_atual,
                        linha_alterada,
                    )

                    # se não for definido multiplas linha no parâmetro
                    if multilinhas is False:
                        # anula nova entrada desse bloco
                        validacao_multilinhas = False

            # escreve o conteúdo no arquivo
            arquivo_aberto.write(conteudo[linha_arquivo] + '\n')

    # retorna o conteúdo do arquivo
    return conteudo


def caminho_existente(caminho):
    """Informa se um arquivo ou pasta existe no caminho indicado.

    Parâmetros:
        caminho: Caminho a ser verificado, relativo ou absoluto.

    Retorna:
        bool: `True` se o caminho existe, `False` caso contrário.

    Exemplos:
        >>> caminho_existente(
        ...     caminho='entrada/base.xlsx',
        ... )
        True
    """
    # Importa recursos do módulo Path
    from pathlib import Path

    # coleta o caminho absoluto do caminho informado
    caminho = coletar_caminho_absoluto(caminho)

    # verifica e retorna se o arquivo existe.
    #   True caso exista e False se não existir.
    return Path(caminho).exists()


def cls():
    """Limpa a tela do terminal.

    Retorna:
        None

    Exemplos:
        >>> cls()
    """

    comando = [
        'powershell',
        '-Command',
        'cls'
    ]

    executar_comando_terminal(comando=comando)


def coletar_arvore_caminho(caminho):
    """Devolve a pasta que contém o caminho informado.

    Parâmetros:
        caminho: Caminho de referência, relativo ou absoluto.

    Retorna:
        str: Caminho absoluto do diretório que contém o item informado.

    Exemplos:
        >>> coletar_arvore_caminho(
        ...     caminho='entrada/base.xlsx',
        ... )
        'C:\\\\projeto\\\\entrada'
    """
    # Importa recursos do módulo Path
    from pathlib import Path

    # coleta o caminho informado no padrão do objeto Path
    caminho_interno = coletar_caminho_absoluto(caminho)

    # coleta o caminho informado no padrão do objeto Path
    caminho_interno = Path(caminho_interno)

    # coleta a árvore do caminho informado
    arvore_caminho = str(caminho_interno.parent)

    # retorna o caminho absoluto coletado
    return arvore_caminho


def coletar_caminho_absoluto(caminho):
    """Converte um caminho relativo em caminho absoluto.

    Parâmetros:
        caminho: Caminho a ser convertido.

    Retorna:
        str: Caminho absoluto correspondente.

    Exemplos:
        >>> coletar_caminho_absoluto(
        ...     caminho='saida',
        ... )
        'C:\\\\projeto\\\\saida'
    """
    # Importa recursos do módulo Path
    from pathlib import Path

    # coleta o caminho informado no padrão do objeto Path
    caminho_interno = Path(caminho)

    # coleta o caminho absoluto do caminho
    caminho_absoluto = str(caminho_interno.absolute())

    # retorna o caminho absoluto coletado
    return caminho_absoluto


def coletar_extensao_arquivo(caminho):
    """Devolve as extensões de um arquivo.

    Parâmetros:
        caminho: Caminho ou nome do arquivo.

    Retorna:
        list[str]: Extensões encontradas, com o ponto, na ordem em que
            aparecem no nome. Lista vazia se não houver extensão.

    Exemplos:
        >>> coletar_extensao_arquivo(
        ...     caminho='base.xlsx',
        ... )
        ['.xlsx']
        >>> coletar_extensao_arquivo(
        ...     caminho='backup.tar.gz',
        ... )
        ['.tar', '.gz']
    """
    # Importa recursos do módulo Path
    from pathlib import Path

    # coleta a extensão do arquivo
    arquivo = Path(caminho).suffixes

    # retorna a extensão coletada
    return arquivo


def coletar_idioma_so():
    """Devolve o idioma configurado na interface do Windows.

    Retorna:
        str: Localidade do usuário no formato `idioma_REGIÃO`.

    Exemplos:
        >>> coletar_idioma_so()
        'pt_BR'
    """
    # Importa recursos do módulo ctypes
    import ctypes

    # Importa recursos do módulo locale
    import locale

    # coleta as informações do kernel do Windows
    windows_dll = ctypes.windll.kernel32

    # coleta o valor do idioma local do sistema no padrão de ID numÃ©rico
    windows_dll.GetUserDefaultUILanguage()

    # coleta o valor do idioma local do sistema no padrão de escrita abreviada
    idioma = locale.windows_locale[windows_dll.GetUserDefaultUILanguage()]

    # retorna o valor de idioma coletado
    return idioma


def coletar_versao_so():
    """Devolve os dados de identificação do sistema operacional.

    Retorna:
        dict: Dicionário com as chaves `sistema` (nome do SO),
            `release` (versão principal), `version` (build detalhado) e
            `machine` (arquitetura do processador).

    Exemplos:
        >>> versao = coletar_versao_so()
        >>> versao['sistema']
        'Windows'
    """
    # Importa recursos do módulo sys
    from platform import machine, release, system, version

    versao_so = {
        "sistema": system(),
        "release": release(),
        "version": version(),
        "machine": machine()
    }

    return versao_so


def coletar_nome_arquivo(caminho):
    """Extrai o nome de um arquivo, sem a pasta e sem a extensão.

    Parâmetros:
        caminho: Caminho ou nome do arquivo.

    Retorna:
        str: Nome do arquivo sem diretório e sem a última extensão.

    Exemplos:
        >>> coletar_nome_arquivo(
        ...     caminho='entrada/base_clientes.xlsx',
        ... )
        'base_clientes'
    """
    # Importa recursos do módulo Path
    from pathlib import Path

    # coleta o nome do arquivo informado
    arquivo = Path(caminho).stem

    # retorna o nome do arquivo
    return arquivo


def coletar_nome_guias_arquivo_excel(arquivo_excel):
    """Lista os nomes das abas de uma planilha do Excel.

    Parâmetros:
        arquivo_excel: Caminho do arquivo Excel a ser inspecionado.

    Retorna:
        list[str]: Nomes das abas, na ordem do arquivo.

    Exemplos:
        >>> coletar_nome_guias_arquivo_excel(
        ...     arquivo_excel='base.xlsx',
        ... )
        ['Clientes', 'Contratos', 'Resumo']
    """

    # Importa recursos do módulo openpyxl
    from openpyxl import load_workbook

    # trata o caminho com o objeto Path
    caminho_excel = coletar_caminho_absoluto(arquivo_excel)

    # abre um arquivo de Excel e coleta o conteúdo
    conteudo_excel = load_workbook(caminho_excel)

    # define um valor padrão e inicial à lista
    lista_guias = []

    # coleta a lista de guias que o arquivo contém
    lista_guias = conteudo_excel.sheetnames

    # retorna a lista coletada
    return lista_guias


def coletar_pid(nome_processo):
    """Localiza os processos em execução cujo nome contenha o texto informado.

    Parâmetros:
        nome_processo: Trecho do nome do executável a procurar
            (ex.: 'excel' ou 'EXCEL.EXE').

    Retorna:
        list[dict]: Um dicionário por processo encontrado, com as chaves
            `pid`, `name` e `create_time`; esta última contém o instante de
            criação como timestamp Unix em segundos. Lista vazia se nenhum
            processo corresponder.

    Exemplos:
        >>> coletar_pid(
        ...     nome_processo='notepad',
        ... )
        [{'pid': 10432, 'name': 'notepad.exe', 'create_time': 1690000000.0}]
    """
    # Importa recursos do módulo os
    import psutil

    # instancia uma lista vazia
    listaProcessos = []
    # para cada processo na lista de processos
    for processo in psutil.process_iter():
        # tenta executar a ação
        try:
            # coleta o PID, o nome, o tempo de início do processo
            informacao_processo = processo.as_dict(
                attrs=['pid', 'name', 'create_time']
            )

            # se existir um processo com o mesmo nome informado
            if nome_processo.lower() in informacao_processo['name'].lower():
                # salva o nome do processo
                listaProcessos.append(informacao_processo)
        # para a lista de erros informados
        except (
            psutil.NoSuchProcess,
            psutil.AccessDenied,
            psutil.ZombieProcess,
        ):
            # ignora os erros
            ...

    # retorna uma lista de dicionários com o nome do processo coletado
    return listaProcessos


def coletar_tamanho(caminho):
    """Devolve o tamanho de um arquivo, em bytes.

    Parâmetros:
        caminho: Caminho do arquivo a ser medido.

    Retorna:
        int: Tamanho do arquivo em bytes.

    Exceções:
        FileNotFoundError: Quando o caminho informado não existe.

    Exemplos:
        >>> coletar_tamanho(
        ...     caminho='saida/relatorio.pdf',
        ... )
        284517
    """
    import os

    caminho_interno = coletar_caminho_absoluto(caminho)

    return os.path.getsize(caminho_interno)


def coletar_versao_arquivo(caminho_arquivo):
    """Lê a versão gravada nas propriedades de um arquivo do Windows.

    Parâmetros:
        caminho_arquivo: Caminho absoluto do arquivo a ser inspecionado.

    Retorna:
        tuple[int, int, int, int]: Versão em quatro partes, na ordem
            maior, menor, build e revisão.

    Exceções:
        OSError: Quando o arquivo não existe ou não possui informação
            de versão gravada.

    Exemplos:
        >>> coletar_versao_arquivo(
        ...     caminho_arquivo=r'C:\\Arquivos\\chrome.exe',
        ... )
        (114, 0, 5735, 199)
    """
    from ctypes import (
        POINTER,
        Structure,
        WinError,
        byref,
        cast,
        pointer,
        sizeof,
        windll,
    )
    from ctypes.wintypes import (
        BOOL,
        CHAR,
        DWORD,
        LPCVOID,
        LPCWSTR,
        LPDWORD,
        LPVOID,
        PUINT,
        UINT,
    )

    GetFileVersionInfoSizeW = windll.version.GetFileVersionInfoSizeW
    GetFileVersionInfoSizeW.restype = DWORD
    GetFileVersionInfoSizeW.argtypes = [LPCWSTR, LPDWORD]
    GetFileVersionInfoSize = GetFileVersionInfoSizeW

    GetFileVersionInfoW = windll.version.GetFileVersionInfoW
    GetFileVersionInfoW.restype = BOOL
    GetFileVersionInfoW.argtypes = [LPCWSTR, DWORD, DWORD, LPVOID]

    VerQueryValueW = windll.version.VerQueryValueW
    VerQueryValueW.restype = BOOL
    VerQueryValueW.argtypes = [LPCVOID, LPCWSTR, POINTER(LPVOID), PUINT]
    VerQueryValue = VerQueryValueW  # alias

    dwLen = GetFileVersionInfoSize(caminho_arquivo, None)
    if not dwLen:
        raise WinError()

    lpData = (CHAR * dwLen)()
    if not GetFileVersionInfoW(caminho_arquivo, 0, sizeof(lpData), lpData):
        raise WinError()

    class VS_FIXEDFILEINFO(Structure):
        _fields_ = [
            ('dwSignature', DWORD),  # will be 0xFEEF04BD
            ('dwStrucVersion', DWORD),
            ('dwFileVersionMS', DWORD),
            ('dwFileVersionLS', DWORD),
            ('dwProductVersionMS', DWORD),
            ('dwProductVersionLS', DWORD),
            ('dwFileFlagsMask', DWORD),
            ('dwFileFlags', DWORD),
            ('dwFileOS', DWORD),
            ('dwFileType', DWORD),
            ('dwFileSubtype', DWORD),
            ('dwFileDateMS', DWORD),
            ('dwFileDateLS', DWORD),
        ]

    uLen = UINT()
    pointer_informacao_arquivo = POINTER(VS_FIXEDFILEINFO)()
    lplpBuffer = cast(pointer(pointer_informacao_arquivo), POINTER(LPVOID))
    if not VerQueryValue(lpData, '\\', lplpBuffer, byref(uLen)):
        raise WinError()

    informacao_arquivo = pointer_informacao_arquivo.contents
    versao = (
        informacao_arquivo.dwFileVersionMS >> 16,
        informacao_arquivo.dwFileVersionMS & 0xFFFF,
        informacao_arquivo.dwFileVersionLS >> 16,
        informacao_arquivo.dwFileVersionLS & 0xFFFF,
    )

    return versao


def compactar(
    caminho: str,
    arquivo_destino: str,
    modo: str = 'w',
):
    """Gera um arquivo `.zip` com todo o conteúdo de uma pasta.

    Parâmetros:
        caminho: Pasta cujo conteúdo será compactado.
        arquivo_destino: Caminho do `.zip` a ser gerado.
        modo: Modo de abertura do ZIP. Aceita `'w'` para recriar o arquivo
            ou `'a'` para acrescentar ao conteúdo existente.

    Retorna:
        bool: `True` quando a operação é concluída.

    Exemplos:
        >>> compactar(
        ...     caminho='saida',
        ...     arquivo_destino='saida.zip',
        ... )
        True
    """
    # Importa recursos do módulo zipfile
    from zipfile import ZipFile

    lista_caminhos = retornar_arquivos_em_pasta(
        caminho=caminho,
        filtro='./**/*',
    )

    with ZipFile(arquivo_destino, modo) as objeto_zip:
        for arquivo in lista_caminhos:
            caminho_interno_zip = arquivo.replace(caminho, '')
            objeto_zip.write(
                filename=arquivo,
                arcname=caminho_interno_zip,
            )

    return True


def converter_pdf_em_imagem(
    arquivo_pdf: str,
    caminho_saida: str,
    alpha: bool = False,
    zoom: float = 1,
    orientacao: int = 0,
):
    """Converte cada página de um PDF em um arquivo de imagem PNG.

    Parâmetros:
        arquivo_pdf: Caminho do PDF a ser convertido.
        caminho_saida: Pasta que receberá as imagens. Deve existir.
        alpha: Aceita `True` para manter o canal de transparência ou
            `False` para aplicar fundo branco.
        zoom: Fator de escala aplicado à página. Aceita número inteiro ou
            decimal; `1` mantém a escala original e valores maiores
            aumentam a resolução.
        orientacao: Rotação aplicada antes de gerar a imagem. Aceita graus
            inteiros; `0`, `90`, `180` e `270` representam as
            orientações ortogonais.

    Retorna:
        bool: `True` quando todas as páginas são convertidas.

    Exceções:
        Exception: Repassa qualquer erro ocorrido na leitura do PDF ou
            na gravação das imagens.

    Exemplos:
        >>> converter_pdf_em_imagem(
        ...     arquivo_pdf='contrato.pdf',
        ...     caminho_saida='imagens',
        ...     zoom=2,
        ... )
        True
    """
    # Importa recursos do módulo Path
    from pathlib import Path

    # Importa recursos do módulo fitz
    import fitz

    try:
        # trata os caminhos com o objeto Path
        caminho_pdf = coletar_caminho_absoluto(arquivo_pdf)
        caminho_saida_img = coletar_caminho_absoluto(caminho_saida)

        # abre um arquivo de PDF e coleta o conteúdo
        conteudo_pdf = fitz.open(caminho_pdf)

        # para cada página
        for indice in range(conteudo_pdf.page_count):
            # coleta a página atual
            pagina = conteudo_pdf[indice]
            # coleta a rotacao
            rotacao = orientacao
            # coleta o zoom do eixo X e eixo Y
            zoom_x = zoom_y = zoom
            # trata o arquivo de saída
            arquivo_img = (
                Path(caminho_saida_img) / f'arquivo_{str(indice+1)}.png'
            )

            # coleta a matriz da página, combinando o zoom e a rotação
            matriciado = fitz.Matrix(zoom_x, zoom_y).prerotate(rotacao)
            # converte a matriz da página em um mapa
            #   de píxel de imagem adicionando fundo
            mapa_pixel = pagina.get_pixmap(matrix=matriciado, alpha=alpha)

            # salva o mapa de píxel em um arquivo de imagem
            mapa_pixel.save(arquivo_img)

        # retorna True caso a operação tenha sucesso
        return True
    except Exception as erro:
        # retorna False caso a operação tenha sucesso
        raise erro


def copiar_arquivo(arquivo, caminho_destino):
    """Copia um arquivo para outro local, preservando seus metadados.

    Parâmetros:
        arquivo: Caminho do arquivo de origem.
        caminho_destino: Pasta de destino ou novo caminho completo do
            arquivo.

    Retorna:
        str: Caminho do arquivo já copiado.

    Exceções:
        FileNotFoundError: Quando a origem ou a pasta de destino não
            existem.

    Exemplos:
        >>> copiar_arquivo(
        ...     arquivo='base.xlsx',
        ...     caminho_destino='backup',
        ... )
        'C:\\\\projeto\\\\backup\\\\base.xlsx'
    """

    # coleta o caminho absoluto do arquivo
    arquivo = coletar_caminho_absoluto(arquivo)

    # Importa recursos do módulo shutil
    from shutil import copy2

    # copia o arquivo para a pasta de destino informado
    caminho_destino = copy2(arquivo, caminho_destino)

    # retorna o caminho de destino
    return str(caminho_destino)


def copiar_pasta(pasta: str, caminho_destino: str):
    """Copia uma pasta inteira, com todo o seu conteúdo, para outro local.

    Parâmetros:
        pasta: Caminho da pasta de origem.
        caminho_destino: Pasta onde a cópia será criada.

    Retorna:
        str: Caminho absoluto da pasta recém-criada no destino.

    Exceções:
        FileExistsError: Quando já existe uma pasta de mesmo nome no destino.

    Exemplos:
        >>> copiar_pasta(
        ...     pasta='saida',
        ...     caminho_destino='backup',
        ... )
        'C:\\\\projeto\\\\backup\\\\saida'
    """

    # Importa recursos do módulo Path
    from pathlib import Path

    # Importa recursos do módulo shutil
    from shutil import copytree

    # trata o caminho de destino com o objeto Path
    caminho_origem = coletar_caminho_absoluto(pasta)
    caminho_destino = coletar_caminho_absoluto(caminho_destino)

    caminho_destino = str(Path(caminho_destino) / Path(caminho_origem).name)

    # copia a pasta para o destino informado
    copytree(
        str(caminho_origem),
        caminho_destino,
    )

    # retorna o caminho de destino com a pasta copiada
    return caminho_destino


def criar_arquivo_texto(
    caminho,
    dado='',
    encoding='utf8',
    em_bytes: bool = False,
):
    # Importa recursos do módulo Path
    """Cria um arquivo com o conteúdo informado, substituindo o anterior.

    Parâmetros:
        caminho: Caminho do arquivo a ser criado.
        dado: Conteúdo a ser gravado. Deve ser `str` no modo texto e
            `bytes` quando `em_bytes` for `True`.
        encoding: Codificação da gravação. Ignorado no modo binário.
        em_bytes: Aceita `True` para gravar em modo binário ou `False`
            para gravar como texto usando `encoding`.

    Retorna:
        bool: `True` quando a operação é concluída.

    Exceções:
        FileNotFoundError: Quando a pasta de destino não existe.

    Exemplos:
        >>> criar_arquivo_texto(
        ...     caminho='saida/log.txt',
        ...     dado='inicio do processo',
        ... )
        True
    """
    from pathlib import Path

    # caso em_bytes não for verdadeiro
    if em_bytes is False:
        # escreve em um arquivo de texto o conteúdo informado
        Path(caminho).write_text(encoding=encoding, data=dado)
    # caso em_bytes for verdadeiro
    else:
        # escreve em um arquivo de texto o conteúdo informado em bytes
        Path(caminho).write_bytes(data=dado)

    # retorna True caso a operação tenha concluída com sucesso
    return True


def criar_pasta(caminho):
    """Cria uma pasta, incluindo os níveis intermediários que faltarem.

    Parâmetros:
        caminho: Caminho da pasta a ser criada.

    Retorna:
        bool: `True` quando a operação é concluída.

    Exceções:
        FileExistsError: Quando a pasta informada já existe.

    Exemplos:
        >>> criar_pasta(
        ...     caminho='saida/2024/janeiro',
        ... )
        True
    """
    # Importa recursos do módulo Path
    from pathlib import Path

    # trata o caminho com o objeto Path
    caminho_interno = Path(caminho)

    # cria a pasta informada, caso necessário cria
    #   a hierarquia anterior à última pasta
    caminho_interno.mkdir(parents=True)

    # retorna True caso a operação tenha concluída com sucesso
    return True


def descompactar(arquivo, caminho_destino, senha_arquivo=None):
    """Extrai todo o conteúdo de um arquivo `.zip` para uma pasta.

    Parâmetros:
        arquivo: Caminho do `.zip` a ser extraído.
        caminho_destino: Pasta que receberá o conteúdo.
        senha_arquivo: Aceita a senha do pacote em `bytes` ou `None` para
            arquivos sem senha.

    Retorna:
        None

    Exceções:
        RuntimeError: Quando o pacote exige senha e ela não foi
            informada ou está incorreta.

    Exemplos:
        >>> descompactar(
        ...     arquivo='lote.zip',
        ...     caminho_destino='entrada',
        ... )
    """
    # Importa recursos do módulo zipfile
    from zipfile import ZipFile

    with ZipFile(file=arquivo, mode='r') as objeto_zip:
        objeto_zip.extractall(path=caminho_destino, pwd=senha_arquivo)


def escrever_em_arquivo(
    arquivo,
    conteudo,
    modo,
    encoding='utf8',
    nova_linha=None,
):
    """Grava texto em um arquivo, sobrescrevendo ou acrescentando ao final.

    Parâmetros:
        arquivo: Caminho do arquivo a ser gravado.
        conteudo: Texto a ser escrito.
        modo: Modo de abertura. Aceita `'w'` para sobrescrever o arquivo ou
            `'a'` para acrescentar ao final.
        encoding: Codificação usada na gravação.
        nova_linha: Terminador acrescentado após o conteúdo. Aceita
            `'\\r'`, `'\\n'` ou `'\\r\\n'`; qualquer outro valor,
            inclusive `None`, não acrescenta nada.

    Retorna:
        None

    Exemplos:
        >>> escrever_em_arquivo(
        ...     arquivo='log.txt',
        ...     conteudo='etapa concluida',
        ...     modo='a',
        ...     nova_linha='\\n',
        ... )
    """
    from pathlib import Path

    caminho_arquivo = Path(arquivo)
    caminho_arquivo = coletar_caminho_absoluto(caminho_arquivo)

    if (nova_linha is None) or (
        (nova_linha is not None) and (nova_linha not in ['\r', '\n', '\r\n'])
    ):
        nova_linha = ''

    # abre o arquivo definindo o modo de edição e o encoding
    with open(
        caminho_arquivo,
        modo,
        encoding=encoding,
    ) as arquivo:
        # escreve efetivamente o conteúdo no arquivo
        arquivo.write(conteudo + nova_linha)

    # fecha o arquivo
    arquivo.close()


def excluir_arquivo(caminho):
    """Exclui definitivamente um arquivo do disco.

    Parâmetros:
        caminho: Caminho do arquivo a ser excluído.

    Retorna:
        bool: `True` quando a operação é concluída.

    Exceções:
        FileNotFoundError: Quando o arquivo não existe.
        PermissionError: Quando o arquivo está aberto em outro programa.

    Exemplos:
        >>> excluir_arquivo(
        ...     caminho='temp/rascunho.txt',
        ... )
        True
    """
    # Importa recursos do módulo Path
    from pathlib import Path

    caminho = coletar_caminho_absoluto(caminho)

    # exclui o arquivo informado
    Path(caminho).unlink()

    # retorna True caso a operação tenha concluída com sucesso
    return True


def excluir_pasta(caminho, vazia: bool = True):
    """Exclui uma pasta, opcionalmente junto com todo o seu conteúdo.

    Parâmetros:
        caminho: Caminho da pasta a ser excluída.
        vazia: Aceita `True` para exigir que a pasta esteja vazia ou
            `False` para remover recursivamente todo o seu conteúdo.

    Retorna:
        bool: `True` quando a operação é concluída.

    Exceções:
        OSError: Quando `vazia` é `True` e a pasta não está vazia.
        FileNotFoundError: Quando a pasta não existe.

    Exemplos:
        >>> excluir_pasta(
        ...     caminho='temp',
        ...     vazia=False,
        ... )
        True
    """

    caminho_interno = coletar_caminho_absoluto(caminho)

    # Se a pasta estiver vazia
    if vazia is True:
        # Importa recursos do módulo Path
        from pathlib import Path

        # exclui a pasta informada
        Path(caminho_interno).rmdir()

        # retorna True caso a operação tenha concluída com sucesso
        return True
    # Se a pasta não estiver vazia
    else:
        # Importa recursos do módulo rmtree
        from shutil import rmtree

        # exclui a pasta informada e o conteúdo contido nela
        rmtree(caminho_interno)

        # retorna True caso a operação tenha concluída com sucesso
        return True


def executar_comando_terminal(
    comando: list[str],
    tempo_limite: int | None = None,
    nova_linha: bool = False,
    diretorio_execucao: str = None,
):
    """Executa um comando do sistema e devolve o resultado da execução.

    Parâmetros:
        comando: Comando e argumentos, um por item da lista
            (ex.: `['ping', '-n', '1', 'localhost']`).
        tempo_limite: Tempo máximo de execução, em segundos. `None`
            aguarda indefinidamente.
        nova_linha: Aceita `True` para devolver as saídas como texto ou
            `False` para devolvê-las como `bytes`.
        diretorio_execucao: Pasta onde o comando será executado.
            `None` usa o diretório de trabalho atual.

    Retorna:
        subprocess.CompletedProcess: Objeto com `returncode`,
            `stdout` e `stderr` da execução.

    Exceções:
        CalledProcessError: Quando o comando retorna código diferente
            de zero.
        TimeoutExpired: Quando o tempo limite é ultrapassado.
        FileNotFoundError: Quando o executável não é encontrado.

    Exemplos:
        >>> resultado = executar_comando_terminal(
        ...     comando=['hostname'],
        ...     nova_linha=True,
        ... )
        >>> resultado.returncode
        0
    """
    # Importa função de caminho atual no terminal
    from os import getcwd
    # Importa recursos do módulo Path
    from pathlib import Path
    # Importa recurso de execução em terminal
    from subprocess import run


    # Valida se foi informado diretório de execução, caso não
    if diretorio_execucao is None:
        # Define diretório atual como padrão
        diretorio_execucao = Path(getcwd())

    # Executa o comando no sistema
    resultado_comando = run(
        args=comando,
        capture_output=True,
        cwd=str(Path(diretorio_execucao).absolute()),
        timeout=tempo_limite,
        check=True,
        text=nova_linha,
        shell=False,
    )

    # Retorna o resultado da execução do comando
    return resultado_comando


def extrair_texto_ocr(arquivo, linguagem, encoding='utf8'):
    """Extrai texto de uma imagem por reconhecimento óptico de caracteres.

    Parâmetros:
        arquivo: Caminho da imagem a ser lida.
        linguagem: Código do idioma do texto, no padrão do Tesseract
            (ex.: 'por' para português, 'eng' para inglês).
        encoding: Codificação usada para decodificar a saída do
            Tesseract.

    Retorna:
        str: Texto reconhecido na imagem.

    Exemplos:
        >>> extrair_texto_ocr(
        ...     arquivo='nota_fiscal.png',
        ...     linguagem='por',
        ... )
        'NOTA FISCAL ELETRONICA\\nNumero: 000123\\n'
    """
    # Importa recursos do módulo subprocess
    import subprocess

    # abre um arquivo de texto e coleta o conteúdo em bytes
    caminho_arquivo = coletar_caminho_absoluto(arquivo)

    # coleta o texto da imagem usando Pytesseract OCR
    texto_extraido = subprocess.run(
        ('pytesseract', '-l', linguagem, caminho_arquivo),
        stdout=subprocess.PIPE,
        encoding=encoding,
    )

    # retorna o texto coletado
    return texto_extraido.stdout


def finalizar_processo(pid: int):
    """Encerra à força o processo com o PID informado.

    Parâmetros:
        pid: Identificador numérico do processo a ser encerrado.

    Retorna:
        bool: `True` se o processo foi localizado e encerrado,
            `False` se nenhum processo ativo possui aquele PID.

    Exemplos:
        >>> finalizar_processo(
        ...     pid=10432,
        ... )
        True
    """
    # Importa recursos do módulo os
    import psutil

    # instancia um dicionário vazio
    listaProcessos = {}
    # para cada processo na lista de processos
    for processo in psutil.process_iter():
        # tenta executar a ação
        try:
            # coleta o PID, o nome, o tempo de início do processo
            informacao_processo = processo.as_dict(
                attrs=['pid', 'name', 'create_time']
            )

            # se existir um processo com o mesmo nome informado
            if pid == informacao_processo['pid']:
                # encerra o processo informado
                processo.kill()

                # retorna true
                return True
        # para a lista de erros informados
        except (
            psutil.NoSuchProcess,
            psutil.AccessDenied,
            psutil.ZombieProcess,
        ):
            # ignora os erros
            ...

    # retorna um dicionário com o nome do processo coletado
    return False


def gravar_log_em_arquivo(
    arquivo,
    conteudo,
    modo,
    encoding='utf8',
    delimitador=';',
    nova_linha='\r\n',
):
    """Grava uma linha de log em formato delimitado, no estilo CSV.

    Parâmetros:
        arquivo: Caminho do arquivo de log.
        conteudo: Campos da linha, em lista ou tupla, na ordem em que
            devem aparecer.
        modo: Modo de abertura. Aceita `'a'` para acrescentar ao log
            existente ou `'w'` para recriar o arquivo.
        encoding: Codificação usada na gravação.
        delimitador: Caractere que separa os campos.
        nova_linha: Terminador acrescentado ao final. Aceita `'\r'`,
            `'\n'` ou `'\r\n'`; outros valores não acrescentam nada.

    Retorna:
        None

    Exemplos:
        >>> gravar_log_em_arquivo(
        ...     arquivo='exec.csv',
        ...     conteudo=['2024-01-10', 'OK', 'lote 1'],
        ...     modo='a',
        ... )
    """

    # transforma todos os argumentos em lista
    if (not isinstance(conteudo, list)) or (not isinstance(conteudo, tuple)):
        conteudo = list(conteudo)

    # define a variavel
    conteudo = delimitador.join(conteudo)

    # abre o arquivo definindo o modo de edição e o encoding
    escrever_em_arquivo(
        arquivo=arquivo,
        conteudo=conteudo,
        modo=modo,
        encoding=encoding,
        nova_linha=nova_linha,
    )


def janela_dialogo(titulo: str, texto: str, estilo: int = 1):
    """Exibe uma caixa de mensagem do Windows e aguarda a resposta.

    Parâmetros:
        titulo: Texto exibido na barra de título da janela.
        texto: Mensagem exibida no corpo da janela.
        estilo: Código que define os botões exibidos. Aceita `0` para OK,
            `1` para OK/Cancelar, `2` para Anular/Repetir/Ignorar, `3`
            para Sim/Não/Cancelar, `4` para Sim/Não, `5` para
            Repetir/Cancelar e `6` para Cancelar/Tentar novamente/Continuar.

    Retorna:
        int: Código do botão acionado (`1` OK, `2` Cancelar,
            `6` Sim, `7` Não).

    Exemplos:
        >>> janela_dialogo(
        ...     titulo='Atenção',
        ...     texto='Deseja continuar?',
        ...     estilo=4,
        ... )
        6
    """
    # Importa recursos do módulo ctypes
    import ctypes

    # cria o objeto de janela conforme os parâmentros informados
    caixa = ctypes.windll.user32.MessageBoxW(0, texto, titulo, estilo)

    # retorna o objeto
    return caixa


def ler_variavel_ambiente(
    arquivo_config='config.ini',
    nome_bloco_config='padrao',
    nome_variavel=None,
    variavel_sistema: bool = False,
    encoding='utf8',
):
    """Lê uma configuração de um arquivo `.ini` ou do ambiente do sistema.

    Parâmetros:
        arquivo_config: Caminho do arquivo `.ini`. Ignorado quando
            `variavel_sistema` é `True`.
        nome_bloco_config: Seção do arquivo `.ini` a ser lida, o nome
            entre colchetes.
        nome_variavel: Chave a ser lida. `None` devolve todas as
            chaves do bloco.
        variavel_sistema: Aceita `True` para ler a variável do sistema
            operacional ou `False` para ler o arquivo de configuração.
        encoding: Codificação usada na leitura do arquivo.

    Retorna:
        str | dict | None: O valor da chave; o bloco completo como
            dicionário quando `nome_variavel` é `None`; ou `None`
            quando a variável de sistema não existe.

    Exceções:
        KeyError: Quando o bloco ou a chave não existem no arquivo.

    Exemplos:
        >>> ler_variavel_ambiente(
        ...     nome_variavel='usuario',
        ... )
        'admin'
        >>> ler_variavel_ambiente(
        ...     nome_variavel='TEMP',
        ...     variavel_sistema=True,
        ... )
        'C:\\\\Users\\\\joao\\\\AppData\\\\Local\\\\Temp'
    """
    # Importa recursos do módulo os
    import os

    # Importa recursos do módulo ConfigParser
    from configparser import ConfigParser

    # se não for variável de sistema
    if not variavel_sistema is True:
        # instancia o objeto de configuração
        config = ConfigParser()
        # Lê o arquivo de configuração
        config.read(arquivo_config, encoding=encoding)

        # se o nome da variável de ambiente foi informada
        if nome_variavel is not None:
            # coleta o dado da variável de ambiente informado
            bloco = dict(config[nome_bloco_config])
            # retorna o valor coletado
            return bloco[nome_variavel]
        # se o nome da variável de ambiente não foi informada
        else:
            # retorna o todos os dados no
            #   bloco de variáveis de ambiente informado
            return dict(config[nome_bloco_config])
    # se for variável de sistema
    else:
        # retorna o valor da variável de sistema solicitado
        return os.environ.get(nome_variavel)


def logar(
    mensagem,
    nivel,
    arquivo=None,
    modo=None,
    encoding=None,
    formatacao=None,
    handlers=None,
):
    """Registra uma mensagem de log no nível de severidade informado.

    Parâmetros:
        mensagem: Texto a ser registrado.
        nivel: Severidade do registro. Aceita `DEBUG`, `INFO`,
            `WARNING`, `ERROR` ou `CRITICAL`, sem diferenciar
            maiúsculas de minúsculas.
        arquivo: Caminho do arquivo de log. `None` escreve no
            console.
        modo: Modo de abertura do arquivo. Aceita `'a'` para acrescentar ou
            `'w'` para recriar o arquivo.
        encoding: Codificação do arquivo de log.
        formatacao: Máscara de formatação da linha, no padrão do
            `logging` (ex.: `'%(asctime)s %(message)s'`).
        handlers: Lista de handlers customizados do `logging`, para
            destinos além do arquivo e do console.

    Retorna:
        tuple[str, str] | str: Tupla com nível e mensagem registrados;
            ou uma mensagem de erro em texto quando o nível informado é
            inválido.

    Exemplos:
        >>> logar(
        ...     mensagem='processo iniciado',
        ...     nivel='INFO',
        ...     arquivo='exec.log',
        ... )
        ('INFO', 'processo iniciado')
    """
    # Importa recursos do módulo logging
    from logging import (
        CRITICAL,
        DEBUG,
        ERROR,
        INFO,
        WARNING,
        basicConfig,
        critical,
        debug,
        error,
        info,
        warning,
    )

    # define um ní­vel de log
    nivel = nivel.upper()

    # define configurações básicas de log
    basicConfig(
        level=nivel,
        filename=arquivo,
        filemode=modo,
        encoding=encoding,
        format=formatacao,
        handlers=handlers,
    )

    # executa comando de logging conforme o nível:
    if nivel == 'DEBUG':
        debug(mensagem)
    elif nivel == 'INFO':
        info(mensagem)
    elif nivel == 'WARNING':
        warning(mensagem)
    elif nivel == 'ERROR':
        error(mensagem)
    elif nivel == 'CRITICAL':
        critical(mensagem)
    # caso o nível não corresponder aos ní­veis padrões de logging
    else:
        # retorna mensagem de parâmetro inválido
        return 'Parâmetro nível inválido. Por favor, informe-o corretamente.'

    # retorna a mensagem e o nível
    return (nivel, mensagem)


def pasta_esta_vazia(caminho):
    """Informa se uma pasta não contém nenhum arquivo ou subpasta.

    Parâmetros:
        caminho: Caminho da pasta a ser verificada.

    Retorna:
        bool: `True` se a pasta existe e está vazia, `False` caso
            contenha algo ou não exista.

    Exemplos:
        >>> pasta_esta_vazia(
        ...     caminho='entrada',
        ... )
        False
    """
    # Importa recursos do módulo Path
    from pathlib import Path

    caminho = coletar_caminho_absoluto(caminho)
    caminho = Path(caminho)

    # se existir o caminho informado
    if caminho_existente(caminho):
        # coleta de forma recursiva o conteúdo
        #   contido no caminho informado caso existir
        lista_arquivos_pastas = list(caminho.glob('**/*'))

        # se não existir conteúdo no caminho informado
        if len(lista_arquivos_pastas) == 0:
            # retorna True informando que a pasta está vazia
            return True

    # retorna False informando que a pasta não está vazia
    return False


def processo_existente(nome_processo):
    """Informa se há algum processo em execução com o nome informado.

    Parâmetros:
        nome_processo: Trecho do nome do executável a procurar.

    Retorna:
        bool: `True` se ao menos um processo corresponde, `False`
            caso contrário.

    Exemplos:
        >>> processo_existente(
        ...     nome_processo='excel',
        ... )
        True
    """
    # Importa recursos do módulo psutil
    import psutil

    # para cada processo na lista de processos
    for processo in psutil.process_iter():
        # tenta executar a ação
        try:
            # verifica se o nome do processo corresponde ao nome informado
            if nome_processo.lower() in processo.name().lower():
                # caso exista retorna True
                return True
        # para a lista de erros informados
        except (
            psutil.NoSuchProcess,
            psutil.AccessDenied,
            psutil.ZombieProcess,
        ):
            # ignora os erros
            ...
    # retorna False caso não encontre processo com o nome informado
    return False


def recortar(caminho_atual, caminho_novo):
    """Move um arquivo ou pasta para outro local.

    Parâmetros:
        caminho_atual: Caminho atual do arquivo ou pasta.
        caminho_novo: Caminho ou pasta de destino.

    Retorna:
        str: Caminho final do item já movido.

    Exceções:
        FileNotFoundError: Quando a origem não existe.

    Exemplos:
        >>> recortar(
        ...     caminho_atual='entrada/base.xlsx',
        ...     caminho_novo='processados',
        ... )
        'processados\\\\base.xlsx'
    """
    # Importa recursos do módulo Path
    from pathlib import Path
    from shutil import move

    # trata o caminho atual com o objeto Path
    caminho_atual = Path(caminho_atual)
    # trata o caminho novo com o objeto Path
    caminho_novo = Path(caminho_novo)

    # modifica o nome conforme informado
    caminho_novo_str = move(caminho_atual, caminho_novo)

    # retorna o objeto Path com o caminho novo
    return caminho_novo_str


def remover_acentos(
    texto,
    normalizacao='NFKD',
):
    """Remove acentos e caracteres especiais de um texto.

    Parâmetros:
        texto: Texto a ser normalizado.
        normalizacao: Forma de normalização Unicode. Aceita `NFC` e `NFD`
            para equivalência canônica ou `NFKC` e `NFKD` para
            equivalência de compatibilidade. `NFKD` é o padrão.

    Retorna:
        str: Texto contendo apenas caracteres ASCII.

    Exemplos:
        >>> remover_acentos(
        ...     texto='João comeu açaí',
        ... )
        'Joao comeu acai'
    """
    # Importa recursos do módulo unicodedata
    # Importa recursos do módulo re
    import re
    import unicodedata

    # separa os caracteres comuns dos especiais
    # '''
    separacao_acentos = unicodedata.normalize(
        normalizacao,
        texto,
    )

    texto_tratado = ''
    for caractere in separacao_acentos:
        if ord(caractere) < 128:
            texto_tratado += ''.join([caractere])
        else:
            ...

    # remove os caracteres especiais
    texto_limpo = re.sub(
        '[\u007E|\u00B4|\u0060|\u005E|\u00A8|\u0301|\u007E|\u005E|\xc2|\xb4|\xe9|\362]',
        '',
        texto_tratado,
    )

    # retorna o texto tratado
    return texto_limpo


def renomear(caminho, nome_atual, novo_nome):
    """Altera o nome de um arquivo ou pasta dentro do mesmo diretório.

    Parâmetros:
        caminho: Pasta onde o item se encontra.
        nome_atual: Nome atual do arquivo ou pasta, com extensão.
        novo_nome: Novo nome a ser aplicado, com extensão.

    Retorna:
        None

    Exceções:
        FileNotFoundError: Quando o item de origem não existe.

    Exemplos:
        >>> renomear(
        ...     caminho='saida',
        ...     nome_atual='relatorio.pdf',
        ...     novo_nome='relatorio_2024.pdf',
        ... )
    """
    # Importa recursos do módulo Path
    from os import rename
    from pathlib import Path

    # trata o caminho informado e o nome atual com o objeto Path
    nome_atual = Path(caminho) / nome_atual

    # trata o caminho informado e o nome novo com o objeto Path
    novo_nome = Path(caminho) / novo_nome

    # altera o nome atual para o nome novo
    novo_nome_str = rename(nome_atual, novo_nome)

    # retorna o caminho com o nome novo
    return novo_nome_str


def retornar_arquivos_em_pasta(caminho, filtro='**/*'):
    """Lista os caminhos dos arquivos e pastas contidos em um diretório.

    Parâmetros:
        caminho: Pasta a ser percorrida.
        filtro: Máscara de busca no padrão `glob`. Aceita, entre outras,
            `'**/*'` para incluir subpastas, `'*'` para limitar ao primeiro
            nível ou `'*.ext'` para filtrar por uma extensão.

    Retorna:
        list[str]: Caminhos encontrados. Lista vazia quando nada
            corresponde ao filtro.

    Exemplos:
        >>> retornar_arquivos_em_pasta(
        ...     caminho='entrada',
        ...     filtro='*.xlsx',
        ... )
        ['entrada\\\\base.xlsx', 'entrada\\\\clientes.xlsx']
    """
    # Importa recursos do módulo Path
    from pathlib import Path

    # coleta de forma recursiva o conteúdo
    #   contido no caminho informado caso existir
    lista_arquivos = list(Path(caminho).glob(filtro))

    # instancia uma lista vazia
    lista_arquivos_str = []

    # para cada arquivo na lista de arquivos
    for arquivo in lista_arquivos:
        # coleta e salva o arquivo em string
        lista_arquivos_str.append(str(arquivo))

    # retorna uma lista dos arquivos coletados
    return lista_arquivos_str


def retornar_data_hora_atual(parametro):
    """Devolve a data e a hora do momento, no formato informado.

    Parâmetros:
        parametro: Máscara de formatação (ex.: `'%d/%m/%Y'` para
            data, `'%Y%m%d_%H%M%S'` para carimbo em nome de arquivo).

    Retorna:
        str: Data e hora atuais formatadas conforme a máscara.

    Exemplos:
        >>> retornar_data_hora_atual(
        ...     parametro='%d/%m/%Y %H:%M',
        ... )
        '10/01/2024 14:35'
    """
    # Importa recursos do módulo datetime
    import datetime

    # retorna dados de data e/ou hora conforme informado pelo parâmetro.
    return datetime.datetime.now().strftime(parametro)


def transformar_arquivo_em_base64(
    caminho_arquivo: str, encoding: str = 'utf8', erros: str = 'ignore'
):
    """Converte um arquivo em uma string codificada em Base64.

    Parâmetros:
        caminho_arquivo: Caminho do arquivo a ser convertido.
        encoding: Codificação usada para converter o resultado em
            texto.
        erros: Política para erros na conversão final. Aceita os handlers do
            método `decode`, como `'strict'`, `'ignore'` e `'replace'`.

    Retorna:
        str: Conteúdo do arquivo codificado em Base64.

    Exceções:
        FileNotFoundError: Quando o arquivo não existe.

    Exemplos:
        >>> arquivo_base64 = transformar_arquivo_em_base64(
        ...     caminho_arquivo='assinatura.png',
        ... )
        >>> arquivo_base64[:16]
        'iVBORw0KGgoAAAAN'
    """
    import base64

    with open(caminho_arquivo, mode='rb') as arquivo:
        conteudo_arquivo = arquivo.read()
        arquivo_base64 = base64.b64encode(conteudo_arquivo).decode(
            encoding=encoding, errors=erros
        )

    return arquivo_base64
